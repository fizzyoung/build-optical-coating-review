from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any

from _common import (
    SCHEMA_VERSION,
    ScriptError,
    normalize_doi,
    normalize_title,
    run_cli,
    split_semicolon,
    utc_now,
    validate_against_schema,
    write_json,
)

ARRAY_FIELDS = {
    "authors",
    "database_sources",
    "materials",
    "substrates",
    "coatings",
    "deposition_routes",
    "interface_strategies",
    "optical_properties",
    "mechanical_properties",
    "environmental_properties",
    "manufacturing_properties",
    "characterization_methods",
    "themes",
    "intended_uses",
}
ALIASES = {
    "source id": "source_id",
    "source_id": "source_id",
    "title": "title",
    "article title": "title",
    "authors": "authors",
    "author": "authors",
    "year": "year",
    "publication year": "year",
    "journal": "journal",
    "publication title": "journal",
    "document type": "document_type",
    "item type": "document_type",
    "doi": "doi",
    "url": "url",
    "language": "language",
    "database sources": "database_sources",
    "database": "database_sources",
    "record version": "record_version",
    "supersedes source id": "supersedes_source_id",
    "verification level": "verification_level",
    "verification date": "verification_date",
    "full text status": "full_text_status",
    "screening decision": "screening_decision",
    "screening reason": "screening_reason",
    "materials": "materials",
    "substrates": "substrates",
    "coatings": "coatings",
    "deposition routes": "deposition_routes",
    "interface strategies": "interface_strategies",
    "optical properties": "optical_properties",
    "mechanical properties": "mechanical_properties",
    "environmental properties": "environmental_properties",
    "manufacturing properties": "manufacturing_properties",
    "characterization methods": "characterization_methods",
    "themes": "themes",
    "research stage": "research_stage",
    "intended uses": "intended_uses",
    "evidence strength": "evidence_strength",
    "notes": "notes",
}
DOCUMENT_TYPES = {
    "journalarticle": "JOURNAL_ARTICLE",
    "journal article": "JOURNAL_ARTICLE",
    "article": "JOURNAL_ARTICLE",
    "jour": "JOURNAL_ARTICLE",
    "review": "REVIEW",
    "conferencepaper": "CONFERENCE_PAPER",
    "conference paper": "CONFERENCE_PAPER",
    "conf": "CONFERENCE_PAPER",
    "bookchapter": "BOOK_CHAPTER",
    "book chapter": "BOOK_CHAPTER",
    "standard": "STANDARD",
    "patent": "PATENT",
    "thesis": "THESIS",
    "preprint": "PREPRINT",
    "report": "REPORT",
}


def parser() -> argparse.ArgumentParser:
    result = argparse.ArgumentParser(description="Import bibliographic exports into Schema-conformant Source Records.")
    result.add_argument("input", type=Path)
    result.add_argument("output", type=Path)
    result.add_argument("--format", choices=["csv", "tsv", "xlsx", "json", "ris", "bibtex", "endnote-xml"])
    result.add_argument("--sheet", help="XLSX sheet name; defaults to Source Registry, Master Literature, or the active sheet")
    result.add_argument("--database", help="Source database name recorded on every imported record")
    result.add_argument("--query-id", help="Executed query identifier recorded in notes")
    result.add_argument("--log", type=Path, help="Import log path; defaults beside output")
    return result


def clean_header(value: Any) -> str:
    text = re.sub(r"[_\-]+", " ", str(value or "").strip().casefold())
    return " ".join(text.split())


def canonical_row(row: dict[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in row.items():
        canonical = ALIASES.get(clean_header(key), clean_header(key).replace(" ", "_"))
        result[canonical] = value
    return result


def read_delimited(path: Path, delimiter: str) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle, delimiter=delimiter)]


def read_xlsx(path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ScriptError("openpyxl is required for XLSX imports", 2) from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    candidates = [sheet_name] if sheet_name else ["Source Registry", "Master Literature"]
    worksheet = next((workbook[name] for name in candidates if name and name in workbook.sheetnames), workbook.active)
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(rows)]
    except StopIteration:
        return []
    return [dict(zip(headers, values)) for values in rows if any(value not in (None, "") for value in values)]


def read_ris(path: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    current: dict[str, list[str]] = {}
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        match = re.match(r"^([A-Z0-9]{2})\s*-\s*(.*)$", line)
        if not match:
            continue
        tag, value = match.groups()
        if tag == "ER":
            records.append(
                {
                    "title": (current.get("TI") or current.get("T1") or [None])[0],
                    "authors": current.get("AU") or current.get("A1") or [],
                    "year": (current.get("PY") or current.get("Y1") or [None])[0],
                    "journal": (current.get("JO") or current.get("JF") or current.get("T2") or [None])[0],
                    "document_type": (current.get("TY") or [None])[0],
                    "doi": (current.get("DO") or [None])[0],
                    "url": (current.get("UR") or [None])[0],
                }
            )
            current = {}
        else:
            current.setdefault(tag, []).append(value.strip())
    return records


def read_bibtex(path: Path) -> list[dict[str, Any]]:
    text = path.read_text(encoding="utf-8-sig")
    starts = list(re.finditer(r"@(\w+)\s*\{", text))
    records: list[dict[str, Any]] = []
    for index, match in enumerate(starts):
        block = text[match.end() : starts[index + 1].start() if index + 1 < len(starts) else len(text)]
        fields = {}
        block = re.sub(r',\s*(?=\w+\s*=)', '\n', block)
        for field in re.finditer(r"(?ms)^\s*(\w+)\s*=\s*(?:\{(.*?)\}|\"(.*?)\")\s*,?", block):
            fields[field.group(1).casefold()] = (field.group(2) if field.group(2) is not None else field.group(3)).strip()
        records.append(
            {
                "title": fields.get("title"),
                "authors": re.split(r"\s+and\s+", fields.get("author", ""), flags=re.IGNORECASE),
                "year": fields.get("year"),
                "journal": fields.get("journal") or fields.get("booktitle"),
                "document_type": match.group(1),
                "doi": fields.get("doi"),
                "url": fields.get("url"),
            }
        )
    return records


def first_text(element: ET.Element, paths: list[str]) -> str | None:
    for path in paths:
        found = element.find(path)
        if found is not None and found.text:
            return found.text.strip()
    return None


def read_endnote_xml(path: Path) -> list[dict[str, Any]]:
    root = ET.parse(path).getroot()
    records = []
    for record in root.findall(".//record"):
        authors = [item.text.strip() for item in record.findall(".//contributors/authors/author") if item.text]
        records.append(
            {
                "title": first_text(record, [".//titles/title", ".//title"]),
                "authors": authors,
                "year": first_text(record, [".//dates/year", ".//year"]),
                "journal": first_text(record, [".//periodical/full-title", ".//secondary-title"]),
                "document_type": first_text(record, [".//ref-type"]),
                "doi": first_text(record, [".//electronic-resource-num", ".//doi"]),
                "url": first_text(record, [".//urls/related-urls/url", ".//url"]),
            }
        )
    return records


def read_json_records(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(payload, dict) and isinstance(payload.get("records"), list):
        payload = payload["records"]
    if not isinstance(payload, list):
        raise ScriptError("JSON import must be an array or contain a records array", 2)
    result = []
    for item in payload:
        data = item.get("data", item) if isinstance(item, dict) else None
        if not isinstance(data, dict):
            raise ScriptError("Every JSON record must be an object", 2)
        if "creators" in data and "authors" not in data:
            authors = []
            for creator in data.get("creators", []):
                name = creator.get("name") or " ".join(filter(None, [creator.get("firstName"), creator.get("lastName")]))
                if name:
                    authors.append(name)
            data = {**data, "authors": authors}
        result.append(data)
    return result


def parse_year(value: Any) -> int | None:
    if value in (None, ""):
        return None
    match = re.search(r"(?:16|17|18|19|20|21)\d{2}", str(value))
    return int(match.group(0)) if match else None


def document_type(value: Any) -> str:
    text = str(value or "").strip()
    if text in {
        "JOURNAL_ARTICLE", "REVIEW", "CONFERENCE_PAPER", "BOOK_CHAPTER", "STANDARD",
        "PATENT", "THESIS", "PREPRINT", "REPORT", "OTHER",
    }:
        return text
    return DOCUMENT_TYPES.get(text.casefold(), "OTHER")


def stable_source_id(row: dict[str, Any]) -> str:
    if row.get("source_id"):
        return str(row["source_id"]).strip().upper()
    identity = normalize_doi(row.get("doi")) or "|".join(
        [normalize_title(row.get("title")), str(parse_year(row.get("year")) or ""), ";".join(split_semicolon(row.get("authors")))]
    )
    if not identity.strip("|"):
        raise ScriptError("Cannot assign Source_ID without a title, DOI, year, or author", 3)
    return "SRC-" + hashlib.sha256(identity.encode("utf-8")).hexdigest()[:12].upper()


def normalize_record(raw: dict[str, Any], database: str | None, query_id: str | None) -> dict[str, Any]:
    row = canonical_row(raw)
    notes = str(row.get("notes") or "").strip() or None
    if query_id:
        notes = "; ".join(filter(None, [notes, f"query_id={query_id}"]))
    result = {
        "schema_version": SCHEMA_VERSION,
        "source_id": stable_source_id(row),
        "title": str(row.get("title") or "").strip() or None,
        "authors": split_semicolon(row.get("authors")),
        "year": parse_year(row.get("year")),
        "journal": str(row.get("journal") or "").strip() or None,
        "document_type": document_type(row.get("document_type")),
        "doi": normalize_doi(row.get("doi")),
        "url": str(row.get("url") or "").strip() or None,
        "language": str(row.get("language") or "").strip() or None,
        "database_sources": split_semicolon(row.get("database_sources")),
        "record_version": str(row.get("record_version") or "1").strip(),
        "supersedes_source_id": str(row.get("supersedes_source_id") or "").strip() or None,
        "verification_level": str(row.get("verification_level") or "V0").strip().upper(),
        "verification_date": str(row.get("verification_date") or "").strip() or None,
        "full_text_status": str(row.get("full_text_status") or "NOT_REQUESTED").strip().upper(),
        "screening_decision": str(row.get("screening_decision") or "UNSCREENED").strip().upper(),
        "screening_reason": str(row.get("screening_reason") or "").strip() or None,
        "materials": split_semicolon(row.get("materials")),
        "substrates": split_semicolon(row.get("substrates")),
        "coatings": split_semicolon(row.get("coatings")),
        "deposition_routes": split_semicolon(row.get("deposition_routes")),
        "interface_strategies": split_semicolon(row.get("interface_strategies")),
        "optical_properties": split_semicolon(row.get("optical_properties")),
        "mechanical_properties": split_semicolon(row.get("mechanical_properties")),
        "environmental_properties": split_semicolon(row.get("environmental_properties")),
        "manufacturing_properties": split_semicolon(row.get("manufacturing_properties")),
        "characterization_methods": split_semicolon(row.get("characterization_methods")),
        "themes": split_semicolon(row.get("themes")),
        "research_stage": str(row.get("research_stage") or "").strip() or None,
        "intended_uses": [value.upper() for value in split_semicolon(row.get("intended_uses"))],
        "evidence_strength": str(row.get("evidence_strength") or "UNASSESSED").strip().upper(),
        "notes": notes,
    }
    if database and database not in result["database_sources"]:
        result["database_sources"].append(database)
    return result


def detect_format(path: Path) -> str:
    suffix = path.suffix.casefold()
    return {
        ".csv": "csv", ".tsv": "tsv", ".xlsx": "xlsx", ".json": "json",
        ".ris": "ris", ".bib": "bibtex", ".xml": "endnote-xml",
    }.get(suffix) or ""


def main() -> int:
    args = parser().parse_args()
    path = args.input.resolve()
    if not path.is_file():
        raise ScriptError(f"Input file does not exist: {path}", 2)
    format_name = args.format or detect_format(path)
    readers = {
        "csv": lambda: read_delimited(path, ","),
        "tsv": lambda: read_delimited(path, "\t"),
        "xlsx": lambda: read_xlsx(path, args.sheet),
        "json": lambda: read_json_records(path),
        "ris": lambda: read_ris(path),
        "bibtex": lambda: read_bibtex(path),
        "endnote-xml": lambda: read_endnote_xml(path),
    }
    if format_name not in readers:
        raise ScriptError("Cannot detect input format; use --format", 2)
    raw_records = readers[format_name]()
    records = [normalize_record(row, args.database, args.query_id) for row in raw_records]
    id_counts: dict[str, int] = {}
    for raw_record, record in zip(raw_records, records):
        if canonical_row(raw_record).get('source_id'):
            continue
        base_id = record['source_id']
        id_counts[base_id] = id_counts.get(base_id, 0) + 1
        if id_counts[base_id] > 1:
            record['source_id'] = f'{base_id}-R{id_counts[base_id]}'
    errors = []
    seen_ids = set()
    for index, record in enumerate(records):
        if record["source_id"] in seen_ids:
            errors.append(f"record[{index}] duplicates Source_ID {record['source_id']}; run deduplicate_records.py after assigning distinct raw IDs")
        seen_ids.add(record["source_id"])
        errors.extend(f"record[{index}] {item}" for item in validate_against_schema(record, "source-record.schema.json"))
    if errors:
        raise ScriptError("Import validation failed: " + " | ".join(errors), 3)
    write_json(args.output, records, overwrite=False)
    log_path = args.log or args.output.with_suffix(".import-log.json")
    log = {
        "status": "OK",
        "input": str(path),
        "input_format": format_name,
        "database": args.database,
        "query_id": args.query_id,
        "imported_records": len(records),
        "output": str(args.output.resolve()),
        "run_at": utc_now(),
        "limitations": ["No external metadata was inferred or verified during import."],
    }
    write_json(log_path, log, overwrite=False)
    print(json.dumps(log, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    run_cli(main)
